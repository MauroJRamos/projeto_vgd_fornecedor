import logging

import openpyxl

from LogFornecedoresDAO import insertLogFornecedor

# Set the logging level to INFO
logging.basicConfig(level=logging.INFO)

# Create a logger object
logger = logging.getLogger(__name__)

def GetDadosPlanilhas(arquivo):

    try:
        #Recebe todos os arquivos extraidos
        workbook = openpyxl.load_workbook(arquivo)

        # Obtenha uma lista de nomes de todas as planilhas no arquivo
        sheets = workbook.sheetnames
        all_data = []  # Lista para armazenar os dados de todas as planilhas
        # worksheet = workbook['Plan1']
        for sheet_name in sheets:            #pegar mais de uma sheet, excel com mais de uma planilhas
            worksheet = workbook[sheet_name]
            #worksheet = workbook['Plan2']
            categoria = "Não informado"
            table_data = []
            #print("SEGUNDA ETAPA")
            # le o valor das celulas que não são nulas e verifia as colunas mescladas
            for row in worksheet.iter_rows():
                row_data = []
                #for cell in row:
                for index, cell in enumerate(row):
                    if index >= 10:  # Limita o número de colunas a 10
                        break
                    if cell.coordinate in worksheet.merged_cells:
                        if cell.value is not None:
                           # print(f'A célula {cell.coordinate} - está mesclada. {cell.value}')
                            categoria = cell.value
                    else:
                        # if cell.value is not None:
                       # print(f'{cell.value} celula {cell.coordinate}', end='\t')
                        row_data.append(cell.value)
                row_data.append(categoria)
                table_data.append(row_data)

            all_data.append(table_data)  # Adiciona os dados da planilha atual à lista

        workbook.close()  # Agora fechamos o arquivo Excel após o loop

        return all_data  # Retorna os dados de todas as planilhas
    except Exception as e:
        logger.error(f"Erro ao ler planilha: {e}")
        error_message = f"Erro ao ler planilha: {e}"
        insertLogFornecedor(error_message, "Leitura planilha")
        return []
