import openpyxl
from Tools.scripts.dutree import display

if __name__ == '__main__':

  # Carregue o arquivo Excel
  workbook = openpyxl.load_workbook('Retorno de Cotação-.xlsx')

  # Obtenha uma lista de nomes de todas as planilhas no arquivo
  sheets = workbook.sheetnames

  for sheet_name in sheets:
      worksheet = workbook[sheet_name]



      # Ler a primeira linha da planilha (linha com os títulos das colunas)
      for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
          col_titles = row

      # Itere sobre as linhas de dados (a partir da segunda linha)
      for row in worksheet.iter_rows(min_row=2, values_only=True):
          data = {}
          for col_index, cell_value in enumerate(row):
            # Verifique se a célula está mesclada
            is_merged = False
            for row in worksheet.iter_rows():
               for cell in row:
                   if cell.coordinate in worksheet.merged_cells:
                       if cell.value is not None:
                           print(f'A célula {cell.coordinate} - está mesclada.')

               if not is_merged:
                   # Use os títulos das colunas da planilha diretamente para acessar os dados
                   col_name = col_titles[col_index]
                   data[col_name] = cell_value

          # Agora você pode acessar os dados com os títulos de coluna exatos da planilha
          print(data)

  # Feche o arquivo Excel
  workbook.close()